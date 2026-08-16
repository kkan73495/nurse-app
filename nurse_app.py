import streamlit as st
import openai
import json
import pandas as pd
import os
import io
import zipfile
from datetime import datetime, date
import base64
import re
import chinese_calendar
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter  # 新增导入

# ============================================
# 智谱 AI 配置（请替换你的 API Key）
# ============================================
client = openai.OpenAI(
    api_key="2c29cf29c588425bbed70019a7d6ce34.YEnCeeeYbWe2gyko",          # ← 替换成你的真实密钥
    base_url="https://open.bigmodel.cn/api/paas/v4"
)

PROFILE_DIR = "护士绩效档案"

# 岗位定义（仅用于参考提示）
WEEKDAY_POSITIONS = ['责1-1', '责1-2', '责1-3', '责1-4', '责1-5', '夜1-1', '夜1-2', '综合1', '综合2', '主班']
HOLIDAY_POSITIONS = ['总责2-1', '责2-2', '责2-3', '责2-4', '综合6']
ALL_POSITIONS = WEEKDAY_POSITIONS + HOLIDAY_POSITIONS

# 岗位层级映射（用于上报表）
POSITION_HIERARCHY = {
    "质控护士": [],
    "责任护士": ['责1-1', '责1-2', '责1-3', '责1-4', '责1-5'],
    "主班": [],
    "教育护士": [],
    "综合": ['综1', '综2'],
    "总责": ['总责2-1'],
    "责护节假日": ['责2-2', '责2-3', '责2-4'],
    "综合节假日": ['综6'],
    "夜班责护": ['夜1-1', '夜1-2'],
    "责护3": [],
    "健康教育门诊": [],
    "专职专科": []
}
# 生成具体岗位列表（按顺序）
SPECIFIC_POSITIONS = []
for category, positions in POSITION_HIERARCHY.items():
    if positions:
        SPECIFIC_POSITIONS.extend(positions)
    else:
        SPECIFIC_POSITIONS.append(category)  # 单独类别直接作为岗位名

def get_positions_for_date(input_date):
    if isinstance(input_date, str):
        input_date = datetime.strptime(input_date, "%Y-%m-%d").date()
    if chinese_calendar.is_holiday(input_date):
        return HOLIDAY_POSITIONS
    else:
        return WEEKDAY_POSITIONS

def default_position_for_date(input_date):
    return get_positions_for_date(input_date)[0]

# ==================== 工具函数 ====================
def extract_images_from_docx(docx_bytes):
    images = []
    with zipfile.ZipFile(io.BytesIO(docx_bytes), 'r') as z:
        for file_name in z.namelist():
            if file_name.startswith('word/media/') and (file_name.endswith('.jpeg') or file_name.endswith('.jpg') or file_name.endswith('.png')):
                images.append(z.read(file_name))
    return images

def parse_image_with_ai(image_bytes):
    base64_image = base64.b64encode(image_bytes).decode('utf-8')
    prompt = """
    你是一位护理部数据录入专家。请仔细分析这张图片（这是每日护理绩效登记表）。
    请提取出图片中**所有当班护士**的以下信息：
    1. 姓名（只提取中文姓名）
    2. 一级护理人数（可以是小数）
    3. 二级护理人数（可以是小数）

    请按以下格式输出（每行一个护士，用逗号分隔）：
    name: 张美丽, level_1: 2.5, level_2: 3
    name: 李护士, level_1: 0, level_2: 0
    不要输出任何解释，只输出数据。
    """
    response = client.chat.completions.create(
        model="glm-4v-flash",
        messages=[
            {
                "role": "user",
                "content": [
                    {"type": "text", "text": prompt},
                    {"type": "image_url", "image_url": {"url": f"data:image/jpeg;base64,{base64_image}"}}
                ]
            }
        ],
        temperature=0.0,
    )
    content = response.choices[0].message.content.strip()
    pattern = r'name["\s]*[:：]\s*["\']*([^"\'，,}\]]+?)["\']*\s*[,，]\s*level_1["\s]*[:：]\s*([\d.]+)\s*[,，]\s*level_2["\s]*[:：]\s*([\d.]+)'
    matches = re.findall(pattern, content, re.IGNORECASE)
    records = []
    for match in matches:
        name = match[0].strip()
        l1 = float(match[1]) if match[1] else 0.0
        l2 = float(match[2]) if match[2] else 0.0
        records.append({"name": name, "level_1": l1, "level_2": l2})
    if records:
        return records
    else:
        st.error(f"无法从返回内容中提取有效数据。原始返回：\n{content[:500]}")
        return []

def save_records_to_excel(records, shift_date, position, base1=None, extra1=None, base2=None, extra2=None):
    if not records:
        return 0
    os.makedirs(PROFILE_DIR, exist_ok=True)
    saved = 0
    for idx, item in enumerate(records):
        name = item.get("name", "").strip()
        if not name:
            continue
        l1_total = float(item.get("level_1", 0.0))
        l2_total = float(item.get("level_2", 0.0))
        def safe_float(val):
            if val is None or (isinstance(val, float) and pd.isna(val)):
                return 0.0
            try:
                return float(val)
            except:
                return 0.0
        if base1 is not None and extra1 is not None:
            b1 = safe_float(base1[idx]) if idx < len(base1) else l1_total
            e1 = safe_float(extra1[idx]) if idx < len(extra1) else 0.0
        else:
            b1 = l1_total
            e1 = 0.0
        if base2 is not None and extra2 is not None:
            b2 = safe_float(base2[idx]) if idx < len(base2) else l2_total
            e2 = safe_float(extra2[idx]) if idx < len(extra2) else 0.0
        else:
            b2 = l2_total
            e2 = 0.0
        l1_total = b1 + e1
        l2_total = b2 + e2
        file_path = os.path.join(PROFILE_DIR, f"{name}.xlsx")
        new_row = pd.DataFrame([{
            "日期": shift_date,
            "岗位": position,
            "一级基础": round(b1, 2),
            "一出": round(e1, 2),
            "一级合计": round(l1_total, 2),
            "二级基础": round(b2, 2),
            "二出": round(e2, 2),
            "二级合计": round(l2_total, 2),
            "录入时间": datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        }])
        if os.path.exists(file_path):
            try:
                df_existing = pd.read_excel(file_path, engine='openpyxl')
                df_updated = pd.concat([df_existing, new_row], ignore_index=True)
            except:
                df_updated = new_row
        else:
            df_updated = new_row
        try:
            df_updated.to_excel(file_path, index=False, engine='openpyxl')
            saved += 1
        except Exception as e:
            st.warning(f"保存 {name} 的记录时出错：{e}")
    return saved

def query_all_stats(start_date=None, end_date=None):
    if not os.path.exists(PROFILE_DIR):
        return pd.DataFrame()
    all_data = []
    for file_name in os.listdir(PROFILE_DIR):
        if not file_name.endswith(".xlsx"):
            continue
        nurse_name = file_name.replace(".xlsx", "")
        file_path = os.path.join(PROFILE_DIR, file_name)
        try:
            df = pd.read_excel(file_path, engine='openpyxl')
        except:
            continue
        if "日期" in df.columns:
            df["日期"] = pd.to_datetime(df["日期"]).dt.date
            if start_date:
                df = df[df["日期"] >= start_date]
            if end_date:
                df = df[df["日期"] <= end_date]
        if not df.empty:
            total_l1 = df["一级合计"].sum() if "一级合计" in df.columns else 0.0
            total_l2 = df["二级合计"].sum() if "二级合计" in df.columns else 0.0
            days = len(df)
            all_data.append({
                "护士姓名": nurse_name,
                "一级护理总数": total_l1,
                "二级护理总数": total_l2,
                "上班天数": days
            })
    return pd.DataFrame(all_data)

def get_records_for_date(target_date):
    if not os.path.exists(PROFILE_DIR):
        return pd.DataFrame()
    records = []
    for file_name in os.listdir(PROFILE_DIR):
        if not file_name.endswith(".xlsx"):
            continue
        nurse_name = file_name.replace(".xlsx", "")
        file_path = os.path.join(PROFILE_DIR, file_name)
        try:
            df = pd.read_excel(file_path, engine='openpyxl')
        except Exception as e:
            st.warning(f"读取 {nurse_name} 的档案时出错（可能损坏）：{e}。跳过该护士。")
            continue
        if "日期" not in df.columns:
            continue
        df["日期"] = pd.to_datetime(df["日期"]).dt.date
        df_day = df[df["日期"] == target_date]
        if not df_day.empty:
            for _, row in df_day.iterrows():
                records.append({
                    "护士": nurse_name,
                    "岗位": row.get("岗位", ""),
                    "一级基础": row.get("一级基础", 0.0),
                    "一出": row.get("一出", 0.0),
                    "一级合计": row.get("一级合计", 0.0),
                    "二级基础": row.get("二级基础", 0.0),
                    "二出": row.get("二出", 0.0),
                    "二级合计": row.get("二级合计", 0.0),
                    "录入时间": row.get("录入时间", "")
                })
    return pd.DataFrame(records)

def update_records_for_date(target_date, df_edited):
    if df_edited.empty:
        delete_records_for_date(target_date)
        return 0
    delete_records_for_date(target_date)
    updated = 0
    for _, row in df_edited.iterrows():
        nurse = row["护士"]
        if pd.isna(nurse) or str(nurse).strip() == "":
            continue
        def safe_float(val):
            if pd.isna(val) or val is None:
                return 0.0
            try:
                return float(val)
            except:
                return 0.0
        position = row["岗位"] if not pd.isna(row["岗位"]) else ""
        b1 = safe_float(row["一级基础"])
        e1 = safe_float(row["一出"])
        l1 = b1 + e1
        b2 = safe_float(row["二级基础"])
        e2 = safe_float(row["二出"])
        l2 = b2 + e2
        save_records_to_excel(
            [{"name": nurse, "level_1": l1, "level_2": l2}],
            target_date.strftime("%Y-%m-%d") if isinstance(target_date, date) else target_date,
            position,
            base1=[b1],
            extra1=[e1],
            base2=[b2],
            extra2=[e2]
        )
        updated += 1
    return updated

def delete_records_for_date(target_date):
    if not os.path.exists(PROFILE_DIR):
        return 0
    deleted = 0
    for file_name in os.listdir(PROFILE_DIR):
        if not file_name.endswith(".xlsx"):
            continue
        file_path = os.path.join(PROFILE_DIR, file_name)
        try:
            df = pd.read_excel(file_path, engine='openpyxl')
        except Exception as e:
            st.warning(f"读取 {file_name} 时出错（可能损坏）：{e}。跳过。")
            continue
        if "日期" not in df.columns:
            continue
        df["日期"] = pd.to_datetime(df["日期"]).dt.date
        mask = df["日期"] == target_date
        if mask.any():
            df = df[~mask]
            try:
                df.to_excel(file_path, index=False, engine='openpyxl')
                deleted += 1
            except Exception as e:
                st.warning(f"保存 {file_name} 时出错：{e}")
    return deleted

# ==================== 生成“上报表”数据 ====================
def generate_upload_report(start_date, end_date):
    if not os.path.exists(PROFILE_DIR):
        return None, None
    nurse_data = {}
    for file_name in os.listdir(PROFILE_DIR):
        if not file_name.endswith(".xlsx"):
            continue
        nurse_name = file_name.replace(".xlsx", "")
        file_path = os.path.join(PROFILE_DIR, file_name)
        try:
            df = pd.read_excel(file_path, engine='openpyxl')
        except:
            continue
        if "日期" not in df.columns:
            continue
        df["日期"] = pd.to_datetime(df["日期"]).dt.date
        if start_date:
            df = df[df["日期"] >= start_date]
        if end_date:
            df = df[df["日期"] <= end_date]
        if df.empty:
            continue
        days_worked = df["日期"].nunique()
        total_l1 = df["一级合计"].sum() if "一级合计" in df.columns else 0.0
        total_l2 = df["二级合计"].sum() if "二级合计" in df.columns else 0.0
        pos_counts = {pos: 0 for pos in SPECIFIC_POSITIONS}
        for pos in df["岗位"]:
            if pos in pos_counts:
                pos_counts[pos] += 1
        row = {
            "护士姓名": nurse_name,
            "工作天数": days_worked,
            "一级合计": total_l1,
            "二级合计": total_l2,
            "特级护理": 0.0,
            "三级护理": 0.0,
            "质量得分": 0.0,
            "加分项": 0.0,
            "满意度": 0.0,
        }
        for pos in SPECIFIC_POSITIONS:
            row[pos] = pos_counts[pos]
        nurse_data[nurse_name] = row
    if not nurse_data:
        return None, None
    df_report = pd.DataFrame.from_dict(nurse_data, orient='index')
    df_report.reset_index(drop=True, inplace=True)
    base_cols = ["护士姓名"]
    work_col = "工作天数"
    volume_cols = ["特级护理", "一级合计", "二级合计", "三级护理"]
    quality_cols = ["质量得分", "加分项", "满意度"]
    pos_cols = SPECIFIC_POSITIONS
    col_order = base_cols + pos_cols + [work_col] + volume_cols + quality_cols
    for col in col_order:
        if col not in df_report.columns:
            df_report[col] = 0
    df_report = df_report[col_order]
    # 构建表头
    header_row1 = []
    header_row2 = []
    header_row3 = []
    header_row1.append("序号")
    header_row2.append("")
    header_row3.append("")
    header_row1.append("姓名")
    header_row2.append("")
    header_row3.append("")
    for pos in pos_cols:
        category = None
        for cat, pos_list in POSITION_HIERARCHY.items():
            if pos in pos_list:
                category = cat
                break
            elif not pos_list and pos == cat:
                category = cat
                break
        if not category:
            category = "其他"
        header_row1.append("护理岗位")
        header_row2.append(category)
        header_row3.append(pos)
    header_row1.append("工作天数")
    header_row2.append("")
    header_row3.append("")
    header_row1.append("工作量")
    header_row2.append("特级护理")
    header_row3.append("")
    header_row1.append("")
    header_row2.append("一级护理")
    header_row3.append("")
    header_row1.append("")
    header_row2.append("二级护理")
    header_row3.append("")
    header_row1.append("")
    header_row2.append("三级护理")
    header_row3.append("")
    header_row1.append("工作质量")
    header_row2.append("质量得分")
    header_row3.append("")
    header_row1.append("")
    header_row2.append("加分项")
    header_row3.append("")
    header_row1.append("满意度")
    header_row2.append("")
    header_row3.append("")
    return df_report, (header_row1, header_row2, header_row3)

# ==================== 导出函数（修复列宽问题） ====================
def export_nurse_details(start_date, end_date):
    if not os.path.exists(PROFILE_DIR):
        return None
    output = io.BytesIO()
    from openpyxl import Workbook
    from openpyxl.styles import Alignment
    wb = Workbook()
    # ---- 上报表 ----
    df_report, headers = generate_upload_report(start_date, end_date)
    ws = wb.active
    ws.title = "上报表"
    if df_report is not None and not df_report.empty:
        h1, h2, h3 = headers
        for col_idx, val in enumerate(h1, 1):
            ws.cell(row=1, column=col_idx, value=val)
        for col_idx, val in enumerate(h2, 1):
            ws.cell(row=2, column=col_idx, value=val)
        for col_idx, val in enumerate(h3, 1):
            ws.cell(row=3, column=col_idx, value=val)
        for row_idx, (_, row) in enumerate(df_report.iterrows(), start=4):
            ws.cell(row=row_idx, column=1, value=row_idx - 3)
            ws.cell(row=row_idx, column=2, value=row["护士姓名"])
            col = 3
            for pos in SPECIFIC_POSITIONS:
                ws.cell(row=row_idx, column=col, value=row[pos])
                col += 1
            ws.cell(row=row_idx, column=col, value=row["工作天数"])
            col += 1
            ws.cell(row=row_idx, column=col, value=row["特级护理"])
            col += 1
            ws.cell(row=row_idx, column=col, value=row["一级合计"])
            col += 1
            ws.cell(row=row_idx, column=col, value=row["二级合计"])
            col += 1
            ws.cell(row=row_idx, column=col, value=row["三级护理"])
            col += 1
            ws.cell(row=row_idx, column=col, value=row["质量得分"])
            col += 1
            ws.cell(row=row_idx, column=col, value=row["加分项"])
            col += 1
            ws.cell(row=row_idx, column=col, value=row["满意度"])
        # 合并单元格
        start_col = 3
        end_col = start_col + len(SPECIFIC_POSITIONS) - 1
        if end_col >= start_col:
            ws.merge_cells(start_row=1, start_column=start_col, end_row=1, end_column=end_col)
            ws.cell(row=1, column=start_col).alignment = Alignment(horizontal='center', vertical='center')
        col = start_col
        prev_cat = None
        cat_start = col
        for pos in SPECIFIC_POSITIONS:
            cat = None
            for cat_name, pos_list in POSITION_HIERARCHY.items():
                if pos in pos_list:
                    cat = cat_name
                    break
                elif not pos_list and pos == cat_name:
                    cat = cat_name
                    break
            if not cat:
                cat = "其他"
            if prev_cat is None:
                prev_cat = cat
                cat_start = col
            elif cat != prev_cat:
                if col - cat_start > 1:
                    ws.merge_cells(start_row=2, start_column=cat_start, end_row=2, end_column=col-1)
                    ws.cell(row=2, column=cat_start).alignment = Alignment(horizontal='center', vertical='center')
                prev_cat = cat
                cat_start = col
            col += 1
        if col - cat_start > 1:
            ws.merge_cells(start_row=2, start_column=cat_start, end_row=2, end_column=col-1)
            ws.cell(row=2, column=cat_start).alignment = Alignment(horizontal='center', vertical='center')
        volume_start_col = start_col + len(SPECIFIC_POSITIONS) + 1
        volume_end_col = volume_start_col + 3
        ws.merge_cells(start_row=1, start_column=volume_start_col, end_row=1, end_column=volume_end_col)
        ws.cell(row=1, column=volume_start_col).alignment = Alignment(horizontal='center', vertical='center')
        quality_start_col = volume_end_col + 1
        quality_end_col = quality_start_col + 1
        ws.merge_cells(start_row=1, start_column=quality_start_col, end_row=1, end_column=quality_end_col)
        ws.cell(row=1, column=quality_start_col).alignment = Alignment(horizontal='center', vertical='center')
        # 修复列宽：使用 get_column_letter
        for col in range(1, ws.max_column + 1):
            ws.column_dimensions[get_column_letter(col)].width = 12
    else:
        ws.cell(row=1, column=1, value="所选时间段内无数据")
    # ---- 汇总 ----
    ws2 = wb.create_sheet("汇总")
    df_stats = query_all_stats(start_date, end_date)
    if not df_stats.empty:
        for r in dataframe_to_rows(df_stats, index=False, header=True):
            ws2.append(r)
    else:
        ws2.append(["所选时间段内无数据"])
    # ---- 个人明细 ----
    for file_name in os.listdir(PROFILE_DIR):
        if not file_name.endswith(".xlsx"):
            continue
        nurse_name = file_name.replace(".xlsx", "")
        file_path = os.path.join(PROFILE_DIR, file_name)
        try:
            df = pd.read_excel(file_path, engine='openpyxl')
        except:
            continue
        if "日期" in df.columns:
            df["日期"] = pd.to_datetime(df["日期"]).dt.date
            if start_date:
                df = df[df["日期"] >= start_date]
            if end_date:
                df = df[df["日期"] <= end_date]
        if df.empty:
            continue
        df_output = df.copy()
        df_output["一级护理显示"] = df_output.apply(
            lambda r: f"{r['一级基础']:.2f}+{r['一出']:.2f}出" if r.get('一出', 0) != 0 else f"{r['一级基础']:.2f}",
            axis=1
        )
        df_output["二级护理显示"] = df_output.apply(
            lambda r: f"{r['二级基础']:.2f}+{r['二出']:.2f}出" if r.get('二出', 0) != 0 else f"{r['二级基础']:.2f}",
            axis=1
        )
        result_df = df_output[["日期", "岗位", "一级护理显示", "一级合计", "二级护理显示", "二级合计"]].copy()
        summary = {
            "日期": "汇总",
            "岗位": "",
            "一级护理显示": "",
            "一级合计": result_df["一级合计"].sum(),
            "二级护理显示": "",
            "二级合计": result_df["二级合计"].sum()
        }
        result_df = pd.concat([result_df, pd.DataFrame([summary])], ignore_index=True)
        sheet_name = nurse_name[:31]
        ws_sheet = wb.create_sheet(sheet_name)
        for r in dataframe_to_rows(result_df, index=False, header=True):
            ws_sheet.append(r)
    wb.save(output)
    output.seek(0)
    return output

# ==================== 辅助函数 ====================
def create_editable_df(records, shift_date):
    positions = get_positions_for_date(shift_date)
    default_pos = positions[0]
    data = []
    for rec in records:
        name = rec.get("name", "")
        l1 = float(rec.get("level_1", 0.0))
        l2 = float(rec.get("level_2", 0.0))
        data.append({
            "姓名": name,
            "岗位": default_pos,
            "一级护理": l1,
            "一出": 0.0,
            "一级合计": l1,
            "二级护理": l2,
            "二出": 0.0,
            "二级合计": l2,
        })
    return pd.DataFrame(data)

# ==================== Streamlit 界面 ====================
st.set_page_config(page_title="护理绩效 · 智能录入", layout="wide")
st.title("📸 护理绩效管理系统")

tab1, tab2, tab3 = st.tabs(["📤 录入数据", "📊 统计汇总", "📁 个人档案"])

# ==================== 录入选项卡 ====================
with tab1:
    st.subheader("📥 录入方式")
    mode = st.radio("选择录入方式", ["🤖 AI 智能识别", "✏️ 手动录入", "📅 按日期编辑记录"], horizontal=True)
    
    if mode == "🤖 AI 智能识别":
        col1, col2 = st.columns([1, 1])
        with col1:
            shift_date = st.date_input("📅 班次日期", datetime.now())
            pos_list = get_positions_for_date(shift_date)
            st.caption(f"当前日期岗位参考：{', '.join(pos_list)}")
            uploaded_file = st.file_uploader("上传照片或 Word 文档", type=["png", "jpg", "jpeg", "docx"])
            image_bytes = None
            if uploaded_file is not None:
                if uploaded_file.type == "application/vnd.openxmlformats-officedocument.wordprocessingml.document" or uploaded_file.name.endswith('.docx'):
                    st.info("📄 正在提取Word中的图片...")
                    file_bytes = uploaded_file.read()
                    images = extract_images_from_docx(file_bytes)
                    if images:
                        image_bytes = images[0]
                        st.image(image_bytes, caption="提取出的表格图片", use_container_width=True)
                    else:
                        st.error("未能提取图片，请直接上传图片文件。")
                else:
                    image_bytes = uploaded_file.read()
                    st.image(image_bytes, caption="上传的图片", use_container_width=True)
            
            if image_bytes and st.button("🚀 开始识别", type="primary"):
                with st.spinner("AI 正在分析图片..."):
                    records = parse_image_with_ai(image_bytes)
                if records:
                    df_edit = create_editable_df(records, shift_date)
                    st.session_state['recognition_df'] = df_edit
                    st.session_state['recognition_date'] = shift_date.strftime("%Y-%m-%d")
                    st.success(f"识别出 {len(records)} 位护士，请在下表中核对并编辑（可新增行）。")
                else:
                    st.warning("未识别到有效数据，请检查图片或尝试手动录入。")
        
        with col2:
            if 'recognition_df' in st.session_state and not st.session_state['recognition_df'].empty:
                df_edit = st.session_state['recognition_df']
                edited_df = st.data_editor(
                    df_edit,
                    column_config={
                        "姓名": st.column_config.TextColumn("姓名", required=True),
                        "岗位": st.column_config.TextColumn("岗位"),
                        "一级护理": st.column_config.NumberColumn("一级护理", min_value=0.0, step=0.01, format="%.2f"),
                        "一出": st.column_config.NumberColumn("一出", min_value=0.0, step=0.01, format="%.2f"),
                        "一级合计": st.column_config.NumberColumn("一级合计", disabled=True, format="%.2f"),
                        "二级护理": st.column_config.NumberColumn("二级护理", min_value=0.0, step=0.01, format="%.2f"),
                        "二出": st.column_config.NumberColumn("二出", min_value=0.0, step=0.01, format="%.2f"),
                        "二级合计": st.column_config.NumberColumn("二级合计", disabled=True, format="%.2f"),
                    },
                    hide_index=True,
                    use_container_width=True,
                    num_rows="dynamic",
                    key="editor"
                )
                if st.button("✅ 确认保存此数据"):
                    records_to_save = []
                    for _, row in edited_df.iterrows():
                        name = row["姓名"].strip()
                        if not name:
                            continue
                        pos = row["岗位"]
                        b1 = float(row["一级护理"])
                        e1 = float(row["一出"])
                        l1 = b1 + e1
                        b2 = float(row["二级护理"])
                        e2 = float(row["二出"])
                        l2 = b2 + e2
                        records_to_save.append({
                            "name": name,
                            "level_1": l1,
                            "level_2": l2,
                            "position": pos,
                            "base1": b1,
                            "extra1": e1,
                            "base2": b2,
                            "extra2": e2
                        })
                    if records_to_save:
                        os.makedirs(PROFILE_DIR, exist_ok=True)
                        saved = 0
                        for item in records_to_save:
                            name = item["name"]
                            pos = item["position"]
                            l1 = item["level_1"]
                            l2 = item["level_2"]
                            b1 = item["base1"]
                            e1 = item["extra1"]
                            b2 = item["base2"]
                            e2 = item["extra2"]
                            file_path = os.path.join(PROFILE_DIR, f"{name}.xlsx")
                            new_row = pd.DataFrame([{
                                "日期": st.session_state['recognition_date'],
                                "岗位": pos,
                                "一级基础": round(b1, 2),
                                "一出": round(e1, 2),
                                "一级合计": round(l1, 2),
                                "二级基础": round(b2, 2),
                                "二出": round(e2, 2),
                                "二级合计": round(l2, 2),
                                "录入时间": datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                            }])
                            if os.path.exists(file_path):
                                try:
                                    df_existing = pd.read_excel(file_path, engine='openpyxl')
                                    df_updated = pd.concat([df_existing, new_row], ignore_index=True)
                                except:
                                    df_updated = new_row
                            else:
                                df_updated = new_row
                            try:
                                df_updated.to_excel(file_path, index=False, engine='openpyxl')
                                saved += 1
                            except Exception as e:
                                st.warning(f"保存 {name} 的记录时出错：{e}")
                        st.success(f"成功保存 {saved} 位护士的数据！")
                        del st.session_state['recognition_df']
                        st.rerun()
                    else:
                        st.error("没有有效记录，请检查姓名是否填写。")
            else:
                st.info("识别结果将显示在此处，您可以增删行、修改数值，合计自动计算。")

    elif mode == "✏️ 手动录入":
        st.subheader("✏️ 手动添加一条记录（支持小数，并可输入出数）")
        with st.form("manual_form"):
            col1, col2 = st.columns(2)
            with col1:
                date_manual = st.date_input("日期", datetime.now())
                name_manual = st.text_input("护士姓名")
                pos_options = get_positions_for_date(date_manual)
                position_manual = st.selectbox("岗位", options=pos_options)
            with col2:
                st.markdown("**一级护理**")
                col_a, col_b = st.columns(2)
                with col_a:
                    l1_base = st.number_input("基础值", min_value=0.0, step=0.01, format="%.2f", value=0.0, key="l1_base")
                with col_b:
                    l1_extra = st.number_input("出数", min_value=0.0, step=0.01, format="%.2f", value=0.0, key="l1_extra")
                l1_total = l1_base + l1_extra
                st.text_input("合计（自动计算）", value=f"{l1_total:.2f}", disabled=True, key="l1_total_display")
                
                st.markdown("**二级护理**")
                col_c, col_d = st.columns(2)
                with col_c:
                    l2_base = st.number_input("基础值", min_value=0.0, step=0.01, format="%.2f", value=0.0, key="l2_base")
                with col_d:
                    l2_extra = st.number_input("出数", min_value=0.0, step=0.01, format="%.2f", value=0.0, key="l2_extra")
                l2_total = l2_base + l2_extra
                st.text_input("合计（自动计算）", value=f"{l2_total:.2f}", disabled=True, key="l2_total_display")
            
            submitted = st.form_submit_button("📥 添加记录")
            if submitted:
                if not name_manual.strip():
                    st.error("请输入护士姓名")
                else:
                    save_records_to_excel(
                        [{"name": name_manual.strip(), "level_1": l1_total, "level_2": l2_total}],
                        date_manual.strftime("%Y-%m-%d"),
                        position_manual,
                        base1=[l1_base],
                        extra1=[l1_extra],
                        base2=[l2_base],
                        extra2=[l2_extra]
                    )
                    st.success(f"已添加 {name_manual} 的记录（一级合计：{l1_total:.2f}，二级合计：{l2_total:.2f}）！")
                    st.rerun()

    else:  # 按日期编辑记录
        st.subheader("📅 查看/编辑某一天的所有护士记录")
        edit_date = st.date_input("选择日期", datetime.now())
        col1, col2 = st.columns(2)
        with col1:
            if st.button("加载该日记录"):
                df_day = get_records_for_date(edit_date)
                if not df_day.empty:
                    st.session_state['day_edit_df'] = df_day
                    st.session_state['day_edit_date'] = edit_date
                    st.success(f"找到 {len(df_day)} 条记录")
                else:
                    st.warning("该日期暂无记录，请先录入。")
        with col2:
            if st.button("🗑️ 删除该日所有记录", type="secondary"):
                if st.checkbox("确认删除？此操作不可撤销！"):
                    deleted = delete_records_for_date(edit_date)
                    if deleted > 0:
                        st.success(f"已删除 {deleted} 位护士在该日的记录。请到“统计汇总”页面点击“刷新统计”查看更新。")
                        if 'day_edit_df' in st.session_state:
                            del st.session_state['day_edit_df']
                        st.rerun()
                    else:
                        st.info("没有记录可删除")

        if 'day_edit_df' in st.session_state:
            pos_list = get_positions_for_date(st.session_state['day_edit_date'])
            st.caption(f"当前日期岗位参考：{', '.join(pos_list)}，请手动输入上述岗位名称")
            df_day = st.session_state['day_edit_df']
            edited_df = st.data_editor(
                df_day,
                column_config={
                    "护士": st.column_config.TextColumn("护士", disabled=True),
                    "岗位": st.column_config.TextColumn("岗位"),
                    "一级基础": st.column_config.NumberColumn("一级基础", min_value=0.0, step=0.01, format="%.2f"),
                    "一出": st.column_config.NumberColumn("一出", min_value=0.0, step=0.01, format="%.2f"),
                    "一级合计": st.column_config.NumberColumn("一级合计", disabled=True, format="%.2f"),
                    "二级基础": st.column_config.NumberColumn("二级基础", min_value=0.0, step=0.01, format="%.2f"),
                    "二出": st.column_config.NumberColumn("二出", min_value=0.0, step=0.01, format="%.2f"),
                    "二级合计": st.column_config.NumberColumn("二级合计", disabled=True, format="%.2f"),
                    "录入时间": st.column_config.TextColumn("录入时间", disabled=True),
                },
                hide_index=True,
                num_rows="dynamic",
                key="day_editor"
            )
            if st.button("💾 保存该日所有修改"):
                updated = update_records_for_date(st.session_state['day_edit_date'], edited_df)
                if updated > 0:
                    st.success(f"成功更新 {updated} 位护士的记录！")
                    del st.session_state['day_edit_df']
                    st.rerun()
                else:
                    st.warning("没有记录被更新，请检查。")

# ==================== 统计选项卡 ====================
with tab2:
    st.subheader("📈 按时间段汇总统计")
    col1, col2 = st.columns(2)
    with col1:
        start = st.date_input("开始日期", datetime.now().replace(day=1))
    with col2:
        end = st.date_input("结束日期", datetime.now())
    
    if st.button("刷新统计", key="stats_btn"):
        df_stats = query_all_stats(start, end)
        if not df_stats.empty:
            df_display = df_stats.copy()
            df_display["一级护理总数"] = df_display["一级护理总数"].round(2)
            df_display["二级护理总数"] = df_display["二级护理总数"].round(2)
            st.dataframe(df_display, use_container_width=True)
            total_l1 = df_stats["一级护理总数"].sum()
            total_l2 = df_stats["二级护理总数"].sum()
            col_a, col_b, col_c = st.columns(3)
            col_a.metric("👩‍⚕️ 护士人数", len(df_stats))
            col_b.metric("🏥 一级护理总次数", f"{total_l1:.2f}")
            col_c.metric("🏥 二级护理总次数", f"{total_l2:.2f}")
            st.bar_chart(df_stats.set_index("护士姓名")[["一级护理总数", "二级护理总数"]])
            
            output = io.BytesIO()
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                df_stats.to_excel(writer, sheet_name="汇总", index=False)
            st.download_button(
                label="📥 导出汇总 Excel",
                data=output.getvalue(),
                file_name=f"护理汇总_{datetime.now().strftime('%Y%m%d')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
        else:
            st.warning("该时间段暂无数据")
    
    st.divider()
    st.subheader("📥 导出详细 Excel（含上报表、汇总、个人明细）")
    if st.button("生成并下载详细 Excel", key="export_excel_btn"):
        with st.spinner("正在生成导出文件，请稍候..."):
            excel_data = export_nurse_details(start, end)
        if excel_data:
            st.download_button(
                label="点击下载 Excel 文件",
                data=excel_data,
                file_name=f"护士明细_{datetime.now().strftime('%Y%m%d')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
        else:
            st.warning("没有数据可导出（所选时间段内无记录）。")

# ==================== 个人档案选项卡 ====================
with tab3:
    st.subheader("📂 查看/编辑每位护士的详细记录")
    if os.path.exists(PROFILE_DIR):
        files = [f for f in os.listdir(PROFILE_DIR) if f.endswith(".xlsx")]
        if files:
            selected = st.selectbox("选择护士", [f.replace(".xlsx", "") for f in files])
            if selected:
                file_path = os.path.join(PROFILE_DIR, f"{selected}.xlsx")
                try:
                    df = pd.read_excel(file_path, engine='openpyxl')
                except Exception as e:
                    st.error(f"读取 {selected} 的档案时出错（文件可能损坏）：{e}。请尝试修复或重新录入。")
                    st.stop()
                if "日期" in df.columns:
                    df["日期"] = pd.to_datetime(df["日期"]).dt.date
                st.dataframe(df, use_container_width=True)
                st.metric("累计一级护理", f"{df['一级合计'].sum():.2f}" if '一级合计' in df.columns else "0.00")
                st.metric("累计二级护理", f"{df['二级合计'].sum():.2f}" if '二级合计' in df.columns else "0.00")
                
                st.subheader("✏️ 编辑该护士的历史记录")
                st.caption("修改任意单元格后点击下方保存按钮，将覆盖原文件。")
                if not df.empty and "日期" in df.columns:
                    ref_date = df["日期"].iloc[0]
                else:
                    ref_date = date.today()
                pos_list = get_positions_for_date(ref_date)
                st.caption(f"岗位参考（根据日期 {ref_date}）：{', '.join(pos_list)}")
                
                edited_df = st.data_editor(
                    df,
                    column_config={
                        "日期": st.column_config.DateColumn("日期"),
                        "岗位": st.column_config.TextColumn("岗位"),
                        "一级基础": st.column_config.NumberColumn("一级基础", min_value=0.0, step=0.01, format="%.2f"),
                        "一出": st.column_config.NumberColumn("一出", min_value=0.0, step=0.01, format="%.2f"),
                        "一级合计": st.column_config.NumberColumn("一级合计", disabled=True, format="%.2f"),
                        "二级基础": st.column_config.NumberColumn("二级基础", min_value=0.0, step=0.01, format="%.2f"),
                        "二出": st.column_config.NumberColumn("二出", min_value=0.0, step=0.01, format="%.2f"),
                        "二级合计": st.column_config.NumberColumn("二级合计", disabled=True, format="%.2f"),
                        "录入时间": st.column_config.DatetimeColumn("录入时间", disabled=True),
                    },
                    hide_index=True,
                    num_rows="dynamic",
                    key="edit_history"
                )
                if st.button("💾 保存修改"):
                    try:
                        edited_df.to_excel(file_path, index=False, engine='openpyxl')
                        st.success("修改已保存！")
                        st.rerun()
                    except Exception as e:
                        st.error(f"保存失败：{e}")
        else:
            st.info("暂未录入任何护士数据")
    else:
        st.info("文件夹尚未创建，请先录入数据")